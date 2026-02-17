import os
import re
import pandas as pd
import openpyxl
import statistics
from openpyxl.styles import Alignment

j = 0
folder_path = r"e2e_data"
pattern = r"MOTION:"   # regex or plain text
pattern2 = r"GPIO_16_IRQ:"   # regex or plain text
pattern3 = r"LED_ON:"   # regex or plain text

def lat2Framable(latency, jitter):
    ret_data = [["LATENCY", "JITTER",]]
    for i in range(len(latency)):
        ret_data += [[latency[i], jitter[i]]]

    min_lat = min(latency)
    max_lat = max(latency)
    avg_lat = sum(latency) / len(latency)
    stdev_lat = statistics.pstdev(latency)

    min_jit = min(jitter[1::])
    max_jit = max(jitter[1::])
    avg_jit = sum(jitter[1::]) / len(jitter[1::])
    stdev_jit = statistics.pstdev(jitter[1::])

    ret_data[0] += ["", "", "MIN", "MAX" ,"AVERAGE", "STDEV"]
    ret_data[1] += ["", "Latency", f'=MIN(ABS(A2:A{1+len(latency)}))', f'=MAX(ABS(A2:A{1+len(latency)}))', f'=AVERAGE(ABS(A2:A{1+len(latency)}))', f'=STDEV.P(ABS(A2:A{1+len(latency)}))']
    ret_data[2] += ["", "Jitter", f'=MIN(ABS(B2:B{1+len(latency)}))', f'=MAX(ABS(B2:B{1+len(latency)}))', f'=AVERAGE(ABS(B2:B{1+len(latency)}))', f'=STDEV.P(ABS(B2:B{1+len(latency)}))']

    return ret_data

while True:
    excel_output = f'output_{j}.xlsx'
    station_meas = f'e2e_station_{j}.txt'
    vehicle_meas = f'e2e_vehicle_{j}.txt'

    station_path = os.path.join(folder_path, station_meas)
    vehicle_path = os.path.join(folder_path, vehicle_meas)

    station_found = os.path.isfile(station_path)
    vehicle_found = os.path.isfile(vehicle_path)

    raw_data = [["VEHICLE_GY_ON", "VEHICLE_GY_OFF", "LED_ON", "STATION_GY_ON", "STATION_GY_OFF", "PHOTOTRANSISTOR"]]
    
    if station_found and vehicle_found:
        print("Files found True")
        # your processing code goes here
        # --- 2. READ FILE LINE BY LINE ---

        gy_data = []
        pt_data = []
        flags = []
        count_pt = 0
        count_gy = 0
        with open(station_path, "r", encoding="utf-8") as f2:
            for line in f2:
                data = []
                if re.search(pattern2, line):
                    count_pt += 1
                    idx = line.find(pattern2)
                    if("MULTIPLE" not in line[idx+len(pattern2):] and "NONE" not in line[idx+len(pattern2):]):
                        line=int(line[idx+len(pattern2):])
                        pt_data += [line]
                    else:
                        flags+=[count_pt]
                else:
                    if re.search(pattern, line):
                        count_gy+=1
                        idx = line.find(pattern)
                        line = line[idx+len(pattern):].split(";")
                        if(count_gy not in flags):
                            data = [int(i) for i in line]
                            gy_data += [data[0:2:]]
        f2.close

        count_gy = 0
        count_led = 0
        vehicle_lines = []
        led_on = []
        with open(vehicle_path, "r", encoding="utf-8") as f:
            for line in f:
                if re.search(pattern, line):
                   count_gy+=1
                   idx = line.find(pattern)
                   line=line[idx+len(pattern):].split(";")
                   if(count_gy not in flags):
                        data = [int(i) for i in line]
                        vehicle_lines += [data]
                else:
                    if re.search(pattern3, line):
                        count_led+=1
                        idx = line.find(pattern3)
                        if(count_led not in flags):
                            line=int(line[idx+len(pattern3):])
                            led_on += [line]
        f.close

        m2m = []
        g2g = []
        e2e = []
        pid_tr = []

        m2m_prev = 0
        g2g_prev = 0
        e2e_prev = 0
        pid_prev = 0

        m2m_jitter = []
        g2g_jitter = []
        e2e_jitter = []
        pid_jitter = []

        for i in range(min(len(gy_data),len(pt_data),len(vehicle_lines))):
            full_data = []

            vec_dat = vehicle_lines[i] + [led_on[i]]
            stat_dat = gy_data[i] + [pt_data[i]]
            
            full_data += vec_dat
            full_data += stat_dat

            m2m += [abs(vec_dat[0] - stat_dat[0])/1e6]
            m2m_jitter += [(m2m[-1] - m2m_prev)]
            m2m_prev = m2m[-1]

            duration_stat = stat_dat[1] - stat_dat[0]
            duration_vec = vec_dat[2] - vec_dat[0]

            pid_tr += [abs(duration_vec - duration_stat)/1e6]
            pid_jitter += [(pid_tr[-1] - pid_prev)]
            pid_prev = pid_tr[-1]

            last_stat = stat_dat[-1]

            g2g += [abs(last_stat - vec_dat[2])/1e6]
            g2g_jitter += [(g2g[-1] - g2g_prev)]
            g2g_prev = g2g[-1]

            e2e += [((last_stat - stat_dat[0])-(vec_dat[2] - vec_dat[0]))/1e6]
            #e2e += [(m2m[-1]+ g2g[-1])]
            e2e_jitter += [(e2e[-1]-e2e_prev)]
            e2e_prev = e2e[-1]

            raw_data += [[f'{i}' for i in full_data]]

        # Save to Excel 
        with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
            df = pd.DataFrame(raw_data[1:], columns=raw_data[0])
            df.to_excel(writer, sheet_name="RAW_DATA", index=False, engine="openpyxl")

            m2m_data = lat2Framable(m2m, m2m_jitter)
            df_m2m = pd.DataFrame(m2m_data[1:], columns=m2m_data[0])
            df_m2m.to_excel(writer, sheet_name="M2M", index=False)

            g2g_data = lat2Framable(g2g, g2g_jitter)
            df_g2g = pd.DataFrame(g2g_data[1:], columns=g2g_data[0])
            df_g2g.to_excel(writer, sheet_name="G2G", index=False)

            e2e_data = lat2Framable(e2e, e2e_jitter)
            df_e2e = pd.DataFrame(e2e_data[1:], columns=e2e_data[0])
            df_e2e.to_excel(writer, sheet_name="E2E", index=False)

            pid_data = lat2Framable(pid_tr, pid_jitter)
            df_pid = pd.DataFrame(pid_data[1:], columns=pid_data[0])
            df_pid.to_excel(writer, sheet_name="PID", index=False)

        # Load the workbook AFTER pandas writes it
        wb = openpyxl.load_workbook(excel_output)

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # --- Center all cells ---
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # --- Auto-adjust column widths ---
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    try:
                        value = str(cell.value)
                        max_length = max(max_length, len(value))
                    except:
                        pass

                ws.column_dimensions[col_letter].width = max_length + 2  # padding

        wb.save(excel_output)



        print(f'Excel file created: {excel_output}')
        j += 1  # move to next pair
    else:
        print("Files found False — stopping")
        break
